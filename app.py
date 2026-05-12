import streamlit as st
import pandas as pd
import json
import re
import io
from bs4 import BeautifulSoup
from PIL import Image
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots

st.set_page_config(page_title="NailVesta 运营周报", page_icon="💅", layout="wide")

# ─────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────
st.markdown("""
<style>
.main{background:#f8f8f6}
.block-container{padding-top:1.5rem}
.metric-card{background:white;border-radius:10px;padding:16px 20px;border:1px solid #e8e8e5;margin-bottom:8px}
.metric-label{font-size:11px;color:#888;text-transform:uppercase;letter-spacing:.06em;margin-bottom:4px}
.metric-value{font-size:26px;font-weight:600;line-height:1}
.metric-sub{font-size:11px;margin-top:5px}
.up{color:#3B6D11}.down{color:#A32D2D}.flat{color:#888}
.section-title{font-size:11px;font-weight:600;color:#666;text-transform:uppercase;letter-spacing:.1em;
  padding-bottom:8px;border-bottom:1px solid #e5e5e5;margin:24px 0 16px}
.insight-box{padding:12px 16px;border-radius:8px;font-size:13px;line-height:1.7;margin:10px 0}
.insight-blue{background:#EBF3FB;border-left:3px solid #378ADD;color:#1a4a70}
.insight-red{background:#FDEEEE;border-left:3px solid #E24B4A;color:#7a1f1f}
.insight-amber{background:#FDF3E3;border-left:3px solid #EF9F27;color:#6b3f09}
.insight-green{background:#EEF6E4;border-left:3px solid #639922;color:#2d5012}
.badge{display:inline-block;font-size:11px;padding:2px 8px;border-radius:10px;font-weight:500;margin:2px}
.badge-red{background:#FCEBEB;color:#791F1F}
.badge-amber{background:#FAEEDA;color:#633806}
.badge-green{background:#EAF3DE;color:#27500A}
.badge-blue{background:#E6F1FB;color:#0C447C}
.summary-pill-row{display:flex;flex-wrap:wrap;gap:8px;margin:12px 0 20px}
.pill-ok{background:#EAF3DE;color:#27500A;border:1px solid #97C459;padding:4px 12px;border-radius:20px;font-size:12px;display:inline-block}
.pill-bad{background:#FCEBEB;color:#791F1F;border:1px solid #F09595;padding:4px 12px;border-radius:20px;font-size:12px;display:inline-block}
.pill-warn{background:#FAEEDA;color:#633806;border:1px solid #EF9F27;padding:4px 12px;border-radius:20px;font-size:12px;display:inline-block}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────
def pct(new, old):
    if old == 0: return None
    return (new / old - 1) * 100

def pct_str(new, old, invert=False):
    v = pct(new, old)
    if v is None: return "—"
    sign = "↑" if v > 0 else "↓"
    css = "up" if (v > 0) != invert else "down"
    return f'<span class="{css}">{sign}{abs(v):.1f}%</span>'

def metric_card(label, value, sub="", direction="neutral"):
    css = {"up": "up", "down": "down", "neutral": "flat"}.get(direction, "flat")
    return f"""
    <div class="metric-card">
      <div class="metric-label">{label}</div>
      <div class="metric-value">{value}</div>
      <div class="metric-sub {css}">{sub}</div>
    </div>"""

def parse_html_report(html_content):
    soup = BeautifulSoup(html_content, "html.parser")
    data = {}
    cards = soup.select(".card")
    for card in cards:
        lbl_el = card.select_one(".lbl")
        val_el = card.select_one(".val")
        if lbl_el and val_el:
            k = lbl_el.get_text(strip=True)
            v = val_el.get_text(strip=True)
            data[k] = v
    tables = {}
    for section in soup.select("section"):
        h2 = section.select_one("h2")
        tbl = section.select_one("table")
        if h2 and tbl:
            title = h2.get_text(strip=True)
            rows = []
            headers = [th.get_text(strip=True) for th in tbl.select("tr th")]
            for tr in tbl.select("tr"):
                tds = [td.get_text(strip=True) for td in tr.select("td")]
                if tds:
                    rows.append(tds)
            tables[title] = {"headers": headers, "rows": rows}
    return data, tables

def to_float(s):
    if s is None: return None
    s = str(s).replace("$","").replace(",","").replace("%","").strip()
    try: return float(s)
    except: return None

def extract_screenshot_metrics(img_file):
    """Screenshot uploaded for reference only; metrics entered manually."""
    return ""

# ─────────────────────────────────────────────
# SIDEBAR – FILE UPLOADS
# ─────────────────────────────────────────────
st.sidebar.image("https://via.placeholder.com/200x50/1a1a1a/ffffff?text=NailVesta", width=180)
st.sidebar.markdown("## 📂 上传数据文件")

auction_file   = st.sidebar.file_uploader("Auction Report (HTML)", type=["html"], key="auction")
cancelled_file = st.sidebar.file_uploader("Cancelled Report (HTML)", type=["html"], key="cancelled")
returned_file  = st.sidebar.file_uploader("Returned Report (HTML)", type=["html"], key="returned")
screenshot_file = st.sidebar.file_uploader("核心指标截图 (PNG/JPG)", type=["png","jpg","jpeg"], key="screenshot")

st.sidebar.markdown("---")
st.sidebar.markdown("### ✏️ 手动输入核心指标")
st.sidebar.caption("（截图 OCR 失败时手动填写）")

period_label = st.sidebar.text_input("本期标签", value="5/1–5/11")
with st.sidebar.expander("本周指标", expanded=True):
    cur_orders   = st.number_input("Orders 订单量", value=1310, step=1)
    cur_sku      = st.number_input("SKU sold 销量", value=1923, step=1)
    cur_attach   = st.number_input("连带率", value=1.47, step=0.01, format="%.2f")
    cur_aov      = st.number_input("AOV ($)", value=43.06, step=0.01, format="%.2f")
    cur_asp      = st.number_input("ASP ($)", value=29.33, step=0.01, format="%.2f")
    cur_ctr      = st.number_input("CTR (%)", value=4.13, step=0.01, format="%.2f")
    cur_cvr      = st.number_input("CVR (%)", value=1.01, step=0.01, format="%.2f")
    cur_atc      = st.number_input("Add-to-Cart (%)", value=4.46, step=0.01, format="%.2f")
    cur_gmv      = st.number_input("总 GMV ($)", value=56404, step=1)

with st.sidebar.expander("上周指标", expanded=False):
    prev_orders  = st.number_input("Orders", value=1580, step=1, key="p_ord")
    prev_sku     = st.number_input("SKU sold", value=2264, step=1, key="p_sku")
    prev_attach  = st.number_input("连带率", value=1.43, step=0.01, format="%.2f", key="p_att")
    prev_aov     = st.number_input("AOV ($)", value=41.15, step=0.01, format="%.2f", key="p_aov")
    prev_asp     = st.number_input("ASP ($)", value=28.72, step=0.01, format="%.2f", key="p_asp")
    prev_ctr     = st.number_input("CTR (%)", value=4.35, step=0.01, format="%.2f", key="p_ctr")
    prev_cvr     = st.number_input("CVR (%)", value=0.94, step=0.01, format="%.2f", key="p_cvr")
    prev_atc     = st.number_input("Add-to-Cart (%)", value=7.38, step=0.01, format="%.2f", key="p_atc")
    prev_gmv     = st.number_input("总 GMV ($)", value=65018, step=1, key="p_gmv")

with st.sidebar.expander("4月周均（基准）", expanded=False):
    apr_orders   = st.number_input("Orders/周", value=1524, step=1, key="a_ord")
    apr_sku      = st.number_input("SKU sold/周", value=2255, step=1, key="a_sku")
    apr_attach   = st.number_input("连带率", value=1.48, step=0.01, format="%.2f", key="a_att")
    apr_aov      = st.number_input("AOV ($)", value=43.52, step=0.01, format="%.2f", key="a_aov")
    apr_asp      = st.number_input("ASP ($)", value=29.42, step=0.01, format="%.2f", key="a_asp")
    apr_gmv      = st.number_input("GMV/周 ($)", value=66350, step=1, key="a_gmv")
    apr_ret_rate = st.number_input("退货率 (%)", value=6.71, step=0.01, format="%.2f", key="a_ret")

with st.sidebar.expander("SKU结构 (本周)", expanded=False):
    sku1_cnt = st.number_input("1件 订单数", value=964, step=1)
    sku1_aov = st.number_input("1件 AOV ($)", value=31.18, format="%.2f")
    sku2_cnt = st.number_input("2件 订单数", value=215, step=1)
    sku2_aov = st.number_input("2件 AOV ($)", value=55.52, format="%.2f")
    sku3_cnt = st.number_input("3件 订单数", value=37, step=1)
    sku3_aov = st.number_input("3件 AOV ($)", value=82.72, format="%.2f")
    sku4_cnt = st.number_input("4件 订单数", value=79, step=1)
    sku4_aov = st.number_input("4件 AOV ($)", value=106.13, format="%.2f")
    sku4p_cnt = st.number_input("4+件 订单数", value=15, step=1)
    sku4p_aov = st.number_input("4+件 AOV ($)", value=197.55, format="%.2f")

with st.sidebar.expander("SKU结构 (上周)", expanded=False):
    p_sku1_pct = st.number_input("1件占比%", value=76.1, format="%.1f", key="ps1")
    p_sku1_aov = st.number_input("1件AOV", value=30.08, format="%.2f", key="pa1")
    p_sku2_pct = st.number_input("2件占比%", value=14.7, format="%.1f", key="ps2")
    p_sku2_aov = st.number_input("2件AOV", value=53.72, format="%.2f", key="pa2")
    p_sku3_pct = st.number_input("3件占比%", value=2.3, format="%.1f", key="ps3")
    p_sku3_aov = st.number_input("3件AOV", value=86.43, format="%.2f", key="pa3")
    p_sku4_pct = st.number_input("4件占比%", value=5.3, format="%.1f", key="ps4")
    p_sku4_aov = st.number_input("4件AOV", value=105.92, format="%.2f", key="pa4")
    p_sku4p_pct = st.number_input("4+件占比%", value=1.5, format="%.1f", key="ps5")
    p_sku4p_aov = st.number_input("4+件AOV", value=183.75, format="%.2f", key="pa5")

run_btn = st.sidebar.button("🚀 生成报告", type="primary", use_container_width=True)

# ─────────────────────────────────────────────
# PARSE HTML FILES
# ─────────────────────────────────────────────
auction_data, auction_tables = {}, {}
cancelled_data, cancelled_tables = {}, {}
returned_data, returned_tables = {}, {}

if auction_file:
    auction_data, auction_tables = parse_html_report(auction_file.read().decode("utf-8"))
if cancelled_file:
    cancelled_data, cancelled_tables = parse_html_report(cancelled_file.read().decode("utf-8"))
if returned_file:
    returned_data, returned_tables = parse_html_report(returned_file.read().decode("utf-8"))

# Derived metrics
total_orders_full = 2603  # from cancelled report Total Orders
if cancelled_data:
    v = to_float(cancelled_data.get("Total Orders","").replace(",",""))
    if v: total_orders_full = int(v)

cancel_cnt = 103
cancel_rate = 4.0
if cancelled_data:
    v = to_float(cancelled_data.get("Cancelled Orders",""))
    if v: cancel_cnt = int(v)

returned_pkgs = 194
ret_rate = 5.62
per_order_ret = 38.89
seller_fault = 94
if returned_data:
    v = to_float(returned_data.get("Returned Packages","").replace(",",""))
    if v: returned_pkgs = int(v)

# Auction metrics
auc_total = 284; auc_cancel = 16; auc_ret = 7
auc_aov = 40.51; auc_ret_aov = 41.99
if auction_data:
    v = to_float(auction_data.get("总 Auction Order 数",""))
    if v: auc_total = int(v)
    v = to_float(auction_data.get("Cancelled 订单",""))
    if v: auc_cancel = int(v)
    v = to_float(auction_data.get("申请退货 Return/Refund",""))
    if v: auc_ret = int(v)
    v = to_float(auction_data.get("有效 Auction 平均 AOV",""))
    if v: auc_aov = v
    v = to_float(auction_data.get("退货 Auction 平均 AOV",""))
    if v: auc_ret_aov = v

auc_cancel_rate = round(auc_cancel / auc_total * 100, 1)
auc_ret_rate = round(auc_ret / auc_total * 100, 1)

sku_total = cur_orders
sku_rows = [
    {"tier":"1件","count":sku1_cnt,"aov":sku1_aov,"pct":round(sku1_cnt/sku_total*100,1),"p_pct":p_sku1_pct,"p_aov":p_sku1_aov,"a_aov":34.16},
    {"tier":"2件","count":sku2_cnt,"aov":sku2_aov,"pct":round(sku2_cnt/sku_total*100,1),"p_pct":p_sku2_pct,"p_aov":p_sku2_aov,"a_aov":59.93},
    {"tier":"3件","count":sku3_cnt,"aov":sku3_aov,"pct":round(sku3_cnt/sku_total*100,1),"p_pct":p_sku3_pct,"p_aov":p_sku3_aov,"a_aov":94.40},
    {"tier":"4件","count":sku4_cnt,"aov":sku4_aov,"pct":round(sku4_cnt/sku_total*100,1),"p_pct":p_sku4_pct,"p_aov":p_sku4_aov,"a_aov":None},
    {"tier":"4+件","count":sku4p_cnt,"aov":sku4p_aov,"pct":round(sku4p_cnt/sku_total*100,1),"p_pct":p_sku4p_pct,"p_aov":p_sku4p_aov,"a_aov":130.11},
]

# ─────────────────────────────────────────────
# MAIN REPORT
# ─────────────────────────────────────────────
st.markdown("# 💅 NailVesta 中台运营分析报告")
st.markdown(f"**统计周期：{period_label}（本周，11天直播口径）** ｜ 对比维度：上周（7天）& 4月整月周均")

# Summary pills
st.markdown(f"""
<div class="summary-pill-row">
  <span class="pill-bad">GMV 周环比 {pct(cur_gmv,prev_gmv):.1f}%</span>
  <span class="pill-ok">AOV ↑{pct(cur_aov,prev_aov):.1f}% · ASP ↑{pct(cur_asp,prev_asp):.1f}%</span>
  <span class="pill-bad">Add-to-Cart ↓39.6%</span>
  <span class="pill-ok">退货率 {ret_rate:.2f}% vs 4月 {apr_ret_rate:.2f}%</span>
  <span class="pill-bad">仓库错包集中反弹</span>
  <span class="pill-ok">Auction 退货率 {auc_ret_rate}%</span>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div class="insight-box insight-blue">
<strong>一句话总结：</strong>本期 GMV 与订单量周环比明显下滑，但质量指标全线改善——AOV ↑4.6%（$41.15→$43.06）、ASP ↑2.1%（$28.72→$29.33）、CVR 突破 1%，显示进店流量的消费意愿与客单价结构均在向好。退货率 5.62% 延续 4 月 W4 后的改善趋势，但仓库执行错误激增（Wrong item ↑160%、Missing ↑200%、Damaged ↑400%）是本期最大风险点，与 4 月月报 W4 后错包率回升直接关联，需立即干预。
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════
# 一、核心指标三期对比
# ══════════════════════════════════════════════
st.markdown('<div class="section-title">一、核心经营指标（三期对比）</div>', unsafe_allow_html=True)

# Metrics row
cols = st.columns(5)
metrics = [
    ("Orders 订单量", f"{cur_orders:,}", pct(cur_orders,prev_orders), True),
    ("SKU sold 销量", f"{cur_sku:,}", pct(cur_sku,prev_sku), True),
    ("AOV 订单均价", f"${cur_aov:.2f}", pct(cur_aov,prev_aov), False),
    ("ASP 客单价", f"${cur_asp:.2f}", pct(cur_asp,prev_asp), False),
    ("连带率", f"{cur_attach:.2f}x", pct(cur_attach,prev_attach), False),
]
for i,(lbl,val,chg,invert) in enumerate(metrics):
    if chg is not None:
        sign = "↑" if chg>0 else "↓"
        color = "down" if (chg>0)==invert else "up"
        sub = f'<span class="{color}">{sign}{abs(chg):.1f}% vs 上周</span>'
    else:
        sub = "—"
    with cols[i]:
        st.markdown(metric_card(lbl, val, sub), unsafe_allow_html=True)

cols2 = st.columns(4)
metrics2 = [
    ("CVR 转化率", f"{cur_cvr:.2f}%", pct(cur_cvr,prev_cvr), False),
    ("CTR 点击率", f"{cur_ctr:.2f}%", pct(cur_ctr,prev_ctr), True),
    ("Add-to-Cart Rate", f"{cur_atc:.2f}%", pct(cur_atc,prev_atc), True),
    ("总 GMV（直播）", f"${cur_gmv:,.0f}", pct(cur_gmv,prev_gmv), True),
]
for i,(lbl,val,chg,invert) in enumerate(metrics2):
    if chg is not None:
        sign = "↑" if chg>0 else "↓"
        color = "down" if (chg>0)==invert else "up"
        sub = f'<span class="{color}">{sign}{abs(chg):.1f}% vs 上周</span>'
    else:
        sub = "—"
    with cols2[i]:
        st.markdown(metric_card(lbl, val, sub), unsafe_allow_html=True)

# Three-period comparison table
st.markdown("#### 三期指标对照表")
df_kpi = pd.DataFrame([
    ["Orders 订单量", f"{apr_orders:,.0f}（全渠道）", f"{prev_orders:,}", f"{pct(prev_orders,apr_orders):+.1f}%", f"{cur_orders:,}", f"{pct(cur_orders,prev_orders):+.1f}%", f"{pct(cur_orders,apr_orders):+.1f}%"],
    ["SKU sold 销量", f"{apr_sku:,.0f}（全渠道）", f"{prev_sku:,}", f"{pct(prev_sku,apr_sku):+.1f}%", f"{cur_sku:,}", f"{pct(cur_sku,prev_sku):+.1f}%", f"{pct(cur_sku,apr_sku):+.1f}%"],
    ["连带率", f"{apr_attach:.2f}x", f"{prev_attach:.2f}x", f"{pct(prev_attach,apr_attach):+.1f}%", f"{cur_attach:.2f}x", f"{pct(cur_attach,prev_attach):+.1f}%", f"{pct(cur_attach,apr_attach):+.1f}%"],
    ["AOV 订单均价", f"${apr_aov:.2f}", f"${prev_aov:.2f}", f"{pct(prev_aov,apr_aov):+.1f}%", f"${cur_aov:.2f}", f"{pct(cur_aov,prev_aov):+.1f}%", f"{pct(cur_aov,apr_aov):+.1f}%"],
    ["ASP 客单价", f"${apr_asp:.2f}", f"${prev_asp:.2f}", f"{pct(prev_asp,apr_asp):+.1f}%", f"${cur_asp:.2f}", f"{pct(cur_asp,prev_asp):+.1f}%", f"{pct(cur_asp,apr_asp):+.1f}%"],
    ["CTR 点击率", "—", f"{prev_ctr:.2f}%", "—", f"{cur_ctr:.2f}%", f"{pct(cur_ctr,prev_ctr):+.1f}%", "—"],
    ["CVR 转化率", "—", f"{prev_cvr:.2f}%", "—", f"{cur_cvr:.2f}%", f"{pct(cur_cvr,prev_cvr):+.1f}%", "—"],
    ["Add-to-Cart Rate", "—", f"{prev_atc:.2f}%", "—", f"{cur_atc:.2f}%", f"{pct(cur_atc,prev_atc):+.1f}%", "—"],
    ["总 GMV（直播）", f"${apr_gmv:,.0f}（全渠道）", f"${prev_gmv:,.0f}", f"{pct(prev_gmv,apr_gmv):+.1f}%", f"${cur_gmv:,.0f}", f"{pct(cur_gmv,prev_gmv):+.1f}%", f"{pct(cur_gmv,apr_gmv):+.1f}%"],
], columns=["指标","4月周均（基准）","上周","vs4月","本周","WoW↑↓","本周vs4月"])
st.dataframe(df_kpi, use_container_width=True, hide_index=True)

# Charts
c1, c2 = st.columns(2)
with c1:
    fig = make_subplots(specs=[[{"secondary_y": True}]])
    x = ["4月周均","上周","本周"]
    fig.add_trace(go.Scatter(x=x,y=[apr_gmv,prev_gmv,cur_gmv],name="GMV ($)",
        line=dict(color="#185FA5",width=2.5),marker=dict(size=8),fill="tozeroy",
        fillcolor="rgba(24,95,165,0.07)"), secondary_y=False)
    fig.add_trace(go.Scatter(x=x,y=[apr_orders,prev_orders,cur_orders],name="订单量",
        line=dict(color="#1D9E75",width=2.5,dash="dot"),marker=dict(size=8)), secondary_y=True)
    fig.update_layout(title="GMV & 订单量趋势",height=280,margin=dict(l=0,r=0,t=35,b=0),
        legend=dict(orientation="h",y=1.1),plot_bgcolor="white",paper_bgcolor="white")
    fig.update_yaxes(title_text="GMV ($)",secondary_y=False,tickprefix="$",gridcolor="#f0f0ee")
    fig.update_yaxes(title_text="订单量",secondary_y=True)
    st.plotly_chart(fig, use_container_width=True)

with c2:
    fig2 = make_subplots(specs=[[{"secondary_y": True}]])
    fig2.add_trace(go.Scatter(x=x,y=[apr_aov,prev_aov,cur_aov],name="AOV",
        line=dict(color="#639922",width=2.5),marker=dict(size=8),fill="tozeroy",
        fillcolor="rgba(99,153,34,0.07)"), secondary_y=False)
    fig2.add_trace(go.Scatter(x=x,y=[apr_asp,prev_asp,cur_asp],name="ASP",
        line=dict(color="#185FA5",width=2.5,dash="dot"),marker=dict(size=8)), secondary_y=False)
    fig2.add_trace(go.Scatter(x=x,y=[apr_attach,prev_attach,cur_attach],name="连带率",
        line=dict(color="#EF9F27",width=2.5,dash="dash"),marker=dict(size=8)), secondary_y=True)
    fig2.update_layout(title="AOV / ASP / 连带率趋势",height=280,margin=dict(l=0,r=0,t=35,b=0),
        legend=dict(orientation="h",y=1.1),plot_bgcolor="white",paper_bgcolor="white")
    fig2.update_yaxes(title_text="金额 ($)",secondary_y=False,tickprefix="$",gridcolor="#f0f0ee")
    fig2.update_yaxes(title_text="连带率",secondary_y=True)
    st.plotly_chart(fig2, use_container_width=True)

ca, cb = st.columns(2)
with ca:
    st.markdown("""
    <div class="insight-box insight-red">
    <strong>📉 量级指标：两连跌</strong><br>
    订单量 ↓17.1%（1,580→1,310），GMV ↓13.2%（$65K→$56K）连续下行。上周已较 4 月周均微降，本周进一步低于 4 月水平（↓14.1%）。含劳动节假期结构影响，但幅度需持续关注。连带率 1.47x 从上周低点 1.43x 回升（↑2.8%），量级下滑主要来自流量端而非多件购买行为恶化。
    </div>
    """, unsafe_allow_html=True)
with cb:
    st.markdown("""
    <div class="insight-box insight-green">
    <strong>📈 质量指标：全面改善</strong><br>
    AOV ↑4.6%（$41.15→$43.06）、ASP ↑2.1%（$28.72→$29.33）同步从上周低点反弹，ASP 基本追平 4 月月均（$29.42）。CVR 突破 1%（0.94%→1.01%，↑7.4%）。但 Add-to-Cart ↓39.6%（7.38%→4.46%）是最大预警——ATC 低位若持续，将在下周直接压制 GMV 天花板。
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════
# 二、SKU 结构
# ══════════════════════════════════════════════
st.markdown('<div class="section-title">二、订单 SKU 结构 & AOV 分层</div>', unsafe_allow_html=True)

df_sku = pd.DataFrame([{
    "SKU/单": r["tier"],
    "本周订单数": r["count"],
    "本周占比": f'{r["pct"]:.1f}%',
    "本周 AOV": f'${r["aov"]:.2f}',
    "上周占比": f'{r["p_pct"]:.1f}%',
    "上周 AOV": f'${r["p_aov"]:.2f}',
    "4月 AOV": f'${r["a_aov"]:.2f}' if r["a_aov"] else "—",
} for r in sku_rows])
st.dataframe(df_sku, use_container_width=True, hide_index=True)

sc1, sc2 = st.columns(2)
with sc1:
    tiers = [r["tier"] for r in sku_rows]
    fig_sku = go.Figure()
    fig_sku.add_trace(go.Bar(name="本周",x=tiers,y=[r["aov"] for r in sku_rows],
        marker_color="#185FA5",text=[f'${v:.0f}' for v in [r["aov"] for r in sku_rows]],textposition="outside"))
    fig_sku.add_trace(go.Bar(name="上周",x=tiers,y=[r["p_aov"] for r in sku_rows],
        marker_color="#9FE1CB",text=[f'${v:.0f}' for v in [r["p_aov"] for r in sku_rows]],textposition="outside"))
    fig_sku.add_trace(go.Bar(name="4月月均",x=tiers,
        y=[r["a_aov"] if r["a_aov"] else 0 for r in sku_rows],
        marker_color="#D3D1C7",text=[f'${v:.0f}' if v else "" for v in [r["a_aov"] for r in sku_rows]],textposition="outside"))
    fig_sku.update_layout(barmode="group",title="各档 AOV 三期对比",height=300,
        margin=dict(l=0,r=0,t=35,b=0),plot_bgcolor="white",paper_bgcolor="white",
        legend=dict(orientation="h",y=1.12),yaxis=dict(tickprefix="$",gridcolor="#f0f0ee"))
    st.plotly_chart(fig_sku, use_container_width=True)

with sc2:
    fig_pct = go.Figure()
    pcts_cur = [r["pct"] for r in sku_rows]
    pcts_prev = [r["p_pct"] for r in sku_rows]
    fig_pct.add_trace(go.Bar(name="本周",x=tiers,y=pcts_cur,marker_color="#185FA5",
        text=[f'{v:.1f}%' for v in pcts_cur],textposition="outside"))
    fig_pct.add_trace(go.Bar(name="上周",x=tiers,y=pcts_prev,marker_color="#9FE1CB",
        text=[f'{v:.1f}%' for v in pcts_prev],textposition="outside"))
    fig_pct.update_layout(barmode="group",title="SKU 档位占比对比",height=300,
        margin=dict(l=0,r=0,t=35,b=0),plot_bgcolor="white",paper_bgcolor="white",
        legend=dict(orientation="h",y=1.12),yaxis=dict(ticksuffix="%",gridcolor="#f0f0ee"))
    st.plotly_chart(fig_pct, use_container_width=True)

st.markdown(f"""
<div class="insight-box insight-green">
<strong>✅ 1件单占比从高点回落：</strong>上周 {p_sku1_pct:.1f}% → 本周 {sku1_cnt/sku_total*100:.1f}%，回归 4 月月均水平（73.8%），是 AOV / ASP 同步回升的核心驱动。2 件单微升至 {sku2_cnt/sku_total*100:.1f}%（上周 {p_sku2_pct:.1f}%）是正向信号，但 AOV ${sku2_aov:.2f} 仍低于满减门槛 $60，活动 2 对该档客群激励有限。4+ 件 AOV ${sku4p_aov:.2f} 创近期新高（上周 ${p_sku4p_aov:.2f}），高价值客群消费意愿极强但基数仅 {sku4p_cnt} 单。
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════
# 三、退换货
# ══════════════════════════════════════════════
st.markdown('<div class="section-title">三、退换货健康度</div>', unsafe_allow_html=True)

ret_cols = st.columns(4)
ret_metrics = [
    ("总退货率", f"{ret_rate:.2f}%", f"4月月均 {apr_ret_rate:.2f}% · 改善 {abs(pct(ret_rate,apr_ret_rate)):.1f}%", "up"),
    ("退货包裹（全渠道）", f"{returned_pkgs}", f"11天 · 周均≈{int(returned_pkgs/11*7)}", "neutral"),
    ("Per Order Return", f"${per_order_ret:.2f}", "4月 $46.33 · ↓16.1%", "up"),
    ("Seller Fault", f"{seller_fault}单", "占退货 48.5% · 仓库预警", "down"),
]
for i,(lbl,val,sub,d) in enumerate(ret_metrics):
    with ret_cols[i]:
        st.markdown(metric_card(lbl,val,sub,d), unsafe_allow_html=True)

st.markdown("""
<div class="insight-box insight-green">
<strong>✅ 退货率持续改善，延续 4 月 W4 趋势：</strong>4 月月内退货率从 W1 7.87% 持续降至 W4 5.26%，本期 5.62% 衔接这一方向，属真实且具延续性的改善。但需警惕：仓库错包激增若不及时遏制，将在未来 2–3 周蚕食这一成果。
</div>
""", unsafe_allow_html=True)

ra, rb = st.columns(2)
with ra:
    st.markdown("#### 退货原因构成")
    ret_reasons = [
        ["No longer needed", 100, "51.5%", "非卖家"],
        ["Missing package", 20, "10.3%", "物流"],
        ["Item doesn't match desc.", 14, "7.2%", "卖家"],
        ["Wrong item was sent", 14, "7.2%", "仓库 ↑160%"],
        ["Damaged item/packaging", 13, "6.7%", "仓库 ↑400%"],
        ["Missing items", 13, "6.7%", "仓库 ↑200%"],
        ["Item arrived too late", 6, "3.1%", "物流"],
        ["Defective item", 5, "2.6%", "产品"],
        ["Missing parts", 4, "2.1%", "仓库"],
    ]
    df_ret = pd.DataFrame(ret_reasons, columns=["退货原因","数量","占比","性质"])
    if returned_tables.get("Return Reason"):
        t = returned_tables["Return Reason"]
        if t["rows"]:
            df_ret_raw = pd.DataFrame(t["rows"], columns=t["headers"][:len(t["rows"][0])])
            st.dataframe(df_ret_raw, use_container_width=True, hide_index=True)
        else:
            st.dataframe(df_ret, use_container_width=True, hide_index=True)
    else:
        st.dataframe(df_ret, use_container_width=True, hide_index=True)

    fig_ret = go.Figure(go.Pie(
        labels=["Non-fault (No longer needed+物流)","Seller Fault (仓库+产品)"],
        values=[51.5,48.5],
        marker_colors=["#B5D4F4","#E24B4A"],
        hole=0.55,
        textinfo="label+percent",
        textfont_size=11,
    ))
    fig_ret.update_layout(height=230,margin=dict(l=0,r=0,t=10,b=0),
        showlegend=False,paper_bgcolor="white")
    st.plotly_chart(fig_ret, use_container_width=True)

    st.markdown("""
    <div class="insight-box insight-red">
    <strong>🔴 仓库执行集中失误：</strong>Wrong item ↑160% / Missing items ↑200% / Damaged ↑400%，三类仓库 Fault 合计约 44 单（22.7%）。与 4 月月报"W4 大促后错包率从 1.2% 回升至 5%"直接关联，必须本周内回溯 5/1–5/11 出库批次。
    </div>
    """, unsafe_allow_html=True)

with rb:
    st.markdown("#### 高退货 Listing & 款式")
    top_ret = [
        ["DreamWear Collection", 38, "18.4%", "🔴 首位"],
        ["Next Gen Collection", 15, "7.8%", "🟡"],
        ["NEW DROP Collection", 15, "6.9%", "🟡"],
        ["Organizer Binder（非甲片）", 10, "5.5%", "🔴 描述偏差"],
        ["Acai Bloom / Ruby Bloom", "各9", "4.1%", "🟡 延续高位"],
        ["Rosé Petal", 8, "3.7%", "🔴 4月最高危"],
        ["Prism Aura", 9, "4.1%", "🟡"],
    ]
    if returned_tables.get("Top5 高退货产品链接"):
        t = returned_tables["Top5 高退货产品链接"]
        df_tret = pd.DataFrame(t["rows"], columns=t["headers"][:len(t["rows"][0])])
        st.dataframe(df_tret, use_container_width=True, hide_index=True)
    else:
        st.dataframe(pd.DataFrame(top_ret,columns=["产品/款式","退货包裹","占比","风险"]),
            use_container_width=True, hide_index=True)

    st.markdown("#### 退货直播归因")
    ret_src = pd.DataFrame([
        ["直播①",19,"9.8%"],["直播②",6,"3.1%"],
        ["直播合计",25,"12.9%（正常）"],["非直播",13,"6.7%"],
        ["Unknown（历史积累）",156,"80.4%"],
    ],columns=["来源","退货包裹","占比"])
    st.dataframe(ret_src, use_container_width=True, hide_index=True)

    st.markdown("#### 取消原因分布")
    cancel_reasons = pd.DataFrame([
        ["Bought by mistake",45,"43.7%"],["No longer needed",28,"27.2%"],
        ["Incorrect shipping address",9,"8.7%"],["Need to change payment",6,"5.8%"],
        ["Customer overdue to pay",5,"4.9%"],["Discount not as expected",4,"3.9%"],
        ["Forgot to apply coupons",3,"2.9%"],
    ],columns=["取消原因","订单数","占比"])
    if cancelled_tables.get("Cancel Reasons"):
        t = cancelled_tables["Cancel Reasons"]
        df_can = pd.DataFrame(t["rows"],columns=t["headers"][:len(t["rows"][0])])
        st.dataframe(df_can, use_container_width=True, hide_index=True)
    else:
        st.dataframe(cancel_reasons, use_container_width=True, hide_index=True)

    st.markdown("""
    <div class="insight-box insight-amber">
    <strong>⚠️ Bought by mistake 43.7%：</strong>直播冲动下单后反悔是主因，62.1% 取消来自直播归因。非直播取消率 5.7% 高于直播 3.3%，Listing 详情信息待优化。Discount not as expected + Forgot coupons 合计 6.8%，活动优惠传达问题延续 4 月判断。
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════
# 四、Auction
# ══════════════════════════════════════════════
st.markdown('<div class="section-title">四、Auction 渠道专项分析</div>', unsafe_allow_html=True)

ac = st.columns(5)
auc_metrics = [
    ("总 Auction 订单", f"{auc_total}", f"全渠道 {total_orders_full:,} 单的 {auc_total/total_orders_full*100:.1f}%", "neutral"),
    ("Auction 退货率", f"{auc_ret_rate}%", f"全渠道 {ret_rate:.2f}% · 低 {abs(pct(auc_ret_rate,ret_rate)):.0f}%", "up"),
    ("Auction 取消率", f"{auc_cancel_rate}%", f"{auc_cancel}单 · 100% 超时未付", "neutral"),
    ("有效 Auction AOV", f"${auc_aov:.2f}", f"排除 {auc_cancel} 单取消", "neutral"),
    ("退货 Auction AOV", f"${auc_ret_aov:.2f}", f"高于均值 +${auc_ret_aov-auc_aov:.2f}", "neutral"),
]
for i,(lbl,val,sub,d) in enumerate(auc_metrics):
    with ac[i]:
        st.markdown(metric_card(lbl,val,sub,d), unsafe_allow_html=True)

aa, ab = st.columns(2)
with aa:
    st.markdown("#### Auction vs 全渠道健康度")
    df_auc_cmp = pd.DataFrame([
        ["退货率", f"{auc_ret_rate}% ✅", f"{ret_rate:.2f}%"],
        ["取消率", f"{auc_cancel_rate}%（全为超时）⚠️", "4.0%（多种原因）"],
        ["Seller Fault 占退货", "25.0% ✅", "48.5% 🔴"],
        ["退货主因", "No longer needed 75%", "No longer needed 51.5%"],
        ["仓库错包渗透", "极低 ✅", "Wrong/Missing 激增 🔴"],
        ["AOV", f"${auc_aov:.2f}", f"${cur_aov:.2f}"],
    ], columns=["指标","Auction 渠道","全渠道"])
    st.dataframe(df_auc_cmp, use_container_width=True, hide_index=True)
    st.markdown("""
    <div class="insight-box insight-green">
    <strong>✅ Auction 是本期最健康的子渠道：</strong>退货率 2.5% 仅为全渠道的 44%，Seller Fault 占比 25% vs 全渠道 48.5%，仓库错包问题在 Auction 渠道几乎未扩散。唯一优化点：16 单超时取消 = 建立直播后 30 分钟催付提醒，零成本挽回。
    </div>
    """, unsafe_allow_html=True)

with ab:
    st.markdown("#### Auction SKU 销售 vs 取消占比")
    auc_sku_lbls = ["SKU 1","SKU 2","SKU 3","SKU 4"]
    auc_sales = [51.8,32.7,13.7,1.8]
    auc_cancel_pct = [31.2,50.0,12.5,6.2]
    fig_auc = go.Figure()
    fig_auc.add_trace(go.Bar(name="销售占比",x=auc_sku_lbls,y=auc_sales,
        marker_color="#185FA5",text=[f'{v}%' for v in auc_sales],textposition="outside"))
    fig_auc.add_trace(go.Bar(name="取消占比",x=auc_sku_lbls,y=auc_cancel_pct,
        marker_color="#E24B4A",text=[f'{v}%' for v in auc_cancel_pct],textposition="outside"))
    fig_auc.update_layout(barmode="group",height=260,margin=dict(l=0,r=0,t=10,b=0),
        plot_bgcolor="white",paper_bgcolor="white",
        legend=dict(orientation="h",y=1.1),
        yaxis=dict(ticksuffix="%",gridcolor="#f0f0ee",range=[0,65]))
    st.plotly_chart(fig_auc, use_container_width=True)
    st.markdown("""
    <div class="insight-box insight-red">
    <strong>🔴 SKU 2 取消严重失衡：</strong>销量占 32.7% 却贡献 50% 取消，隐含取消率 8.6%（SKU 1 仅 3.4%）。全部取消均为超时未付，需复查 SKU 2 直播展示方式与定价清晰度。
    </div>
    """, unsafe_allow_html=True)

# Auction detail tables
auc_detail_a, auc_detail_b = st.columns(2)
with auc_detail_a:
    st.markdown("**取消产品集中度**")
    df_auc_can = pd.DataFrame([
        ["Choose 2 Sets",12,"75.0%"],
        ["Pick your two sets and size",2,"12.5%"],
        ["其他",2,"12.5%"],
    ],columns=["产品","取消单","占比"])
    if auction_tables.get("Auction Cancel 产品链接"):
        t = auction_tables["Auction Cancel 产品链接"]
        df_auc_can = pd.DataFrame(t["rows"],columns=t["headers"][:len(t["rows"][0])])
    st.dataframe(df_auc_can, use_container_width=True, hide_index=True)
    st.markdown("""<div class="insight-box insight-amber">
    全部 16 单取消均为 <strong>Customer overdue to pay</strong>（超时未付，系统自动取消），非买家主动取消。建立直播结束后 30 分钟催付提醒可零成本挽回。
    </div>""", unsafe_allow_html=True)

with auc_detail_b:
    st.markdown("**退货原因**")
    df_auc_ret = pd.DataFrame([
        ["No longer needed",9,"75.0%"],["Missing package",3,"25.0%"],
    ],columns=["退货原因","数量","占比"])
    if auction_tables.get("Auction Return Reason"):
        t = auction_tables["Auction Return Reason"]
        df_auc_ret = pd.DataFrame(t["rows"],columns=t["headers"][:len(t["rows"][0])])
    st.dataframe(df_auc_ret, use_container_width=True, hide_index=True)
    st.markdown("**退货核心指标**")
    df_auc_ri = pd.DataFrame([
        ["Seller Fault","3","25.0%"],["Request Cancelled","4","33.3%"],
        ["已寄出退回","5","41.7%"],["Refund Only","3","25.0%"],
    ],columns=["指标","数量","占比"])
    st.dataframe(df_auc_ri, use_container_width=True, hide_index=True)

# ══════════════════════════════════════════════
# 五、行动建议
# ══════════════════════════════════════════════
st.markdown('<div class="section-title">五、行动建议（衔接 4 月月报）</div>', unsafe_allow_html=True)
act_a, act_b, act_c = st.columns(3)

with act_a:
    st.markdown("""
    <div class="insight-box insight-red">
    <strong>🔴 立即处理（本周内）</strong><br><br>
    <b>1. 仓库批次核查</b><br>Wrong item ↑160% / Missing ↑200% / Damaged ↑400%，回溯 5/1–5/11 出库批次，锁定拣货问题节点。4 月月报已预警 W4 后错包率升至 5%，本期是直接延续。<br><br>
    <b>2. DreamWear Listing 专查</b><br>退货 38 包裹占 18.4%，核查库存映射与主图实物一致性。<br><br>
    <b>3. Rosé Petal 供应商复盘</b><br>4 月月报"持续最高危款"，本期再入 Top。提取退货评论，决定是否暂停销售。<br><br>
    <b>4. Auction 催付机制</b><br>16 单 100% 超时取消，直播结束后 30 分钟内建立催付提醒，零成本挽回。
    </div>
    """, unsafe_allow_html=True)

with act_b:
    st.markdown("""
    <div class="insight-box insight-amber">
    <strong>🟡 本月跟进</strong><br><br>
    <b>5. Add-to-Cart 归因分析</b><br>↓39.6% 是最大流量预警，复盘直播素材、封面与商品卡，防止持续低位压制下周 GMV。<br><br>
    <b>6. 结账页 1→2 件转化提示</b><br>对 AOV $30–$55 用户推"再加一件立减 $8"，是 AOV / ASP 持续提升的最直接杠杆，4 月月报建议已提，尚未落地。<br><br>
    <b>7. Organizer Binder 描述核查</b><br>非甲片类退货 10 包裹，检查详情页内容物描述与实物一致性。<br><br>
    <b>8. Auction SKU 2 复查</b><br>取消率 8.6% vs SKU 1 的 3.4%，检查直播定价展示与买家预期一致性。
    </div>
    """, unsafe_allow_html=True)

with act_c:
    st.markdown("""
    <div class="insight-box insight-green">
    <strong>🟢 策略层（5 月持续）</strong><br><br>
    <b>9. 活动 2 发券精准化</b><br>活动 2 延续至 5/4，约 80% 领券为无效领取，限制资格为历史多件购买用户，目标使用率从 9% 提升至 30%+，ROI 从 10.97x 恢复至 15x+。<br><br>
    <b>10. 非直播 Listing 优化</b><br>非直播取消率 5.7% 持续高于直播 3.3%，完善主图尺码信息减少搜索流量预期差。<br><br>
    <b>11. 统一 GMV 统计口径</b><br>直播口径与全渠道口径混用影响月度趋势判断，本月内确定统一标准。<br><br>
    <b>12. 连带率 & 错包率预警线监控</b><br>连带率预警线 1.5x（本期 1.47x 已低于此），错包率预警线 2%（本期已突破）。
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════
def build_excel():
    wb = openpyxl.Workbook()

    # Colors
    NAVY   = "1F3864"
    BLUE   = "2E75B6"
    LBLUE  = "DEEAF1"
    GREEN  = "375623"
    LGREEN = "E2EFDA"
    RED    = "7B0000"
    LRED   = "FCE4D6"
    AMBER  = "7F6000"
    LAMBER = "FFF2CC"
    WHITE  = "FFFFFF"
    LGRAY  = "F2F2F2"
    MID    = "D9D9D9"

    def hdr_font(color=WHITE, bold=True, size=11):
        return Font(name="Arial", color=color, bold=bold, size=size)

    def cell_font(bold=False, size=10, color="000000"):
        return Font(name="Arial", bold=bold, size=size, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def border(style="thin"):
        s = Side(style=style, color="BFBFBF")
        return Border(left=s, right=s, top=s, bottom=s)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left_al():
        return Alignment(horizontal="left", vertical="center", wrap_text=True)

    def write_section_title(ws, row, col, text, colspan, bg=NAVY):
        ws.cell(row=row, column=col, value=text).font = hdr_font(WHITE, True, 12)
        ws.cell(row=row, column=col).fill = fill(bg)
        ws.cell(row=row, column=col).alignment = center()
        ws.cell(row=row, column=col).border = border()
        if colspan > 1:
            ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col+colspan-1)

    def write_header_row(ws, row, headers, widths=None, start_col=1, bg=BLUE):
        for i, h in enumerate(headers):
            c = ws.cell(row=row, column=start_col+i, value=h)
            c.font = hdr_font(); c.fill = fill(bg); c.alignment = center(); c.border = border()

    def write_data_row(ws, row, values, start_col=1, alt=False, bold=False):
        bg = LGRAY if alt else WHITE
        for i, v in enumerate(values):
            c = ws.cell(row=row, column=start_col+i, value=v)
            c.font = cell_font(bold=bold); c.fill = fill(bg)
            c.alignment = center(); c.border = border()

    def write_metric_block(ws, start_row, start_col, label, value, sub, val_color=None):
        ws.cell(row=start_row, column=start_col, value=label).font = Font(name="Arial",size=9,color="666666")
        ws.cell(row=start_row, column=start_col).fill = fill(LGRAY)
        ws.cell(row=start_row, column=start_col).alignment = center()
        vc = ws.cell(row=start_row+1, column=start_col, value=value)
        vc.font = Font(name="Arial", size=14, bold=True, color=val_color or "1F3864")
        vc.fill = fill(WHITE); vc.alignment = center()
        ws.cell(row=start_row+2, column=start_col, value=sub).font = Font(name="Arial",size=8,color="888888")
        ws.cell(row=start_row+2, column=start_col).fill = fill(WHITE)
        ws.cell(row=start_row+2, column=start_col).alignment = center()
        for r in range(start_row, start_row+3):
            ws.cell(row=r, column=start_col).border = border()

    # ── SHEET 1: 核心指标 ──────────────────────────────
    ws1 = wb.active
    ws1.title = "核心指标"
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 22
    for col in ["B","C","D","E","F","G"]:
        ws1.column_dimensions[col].width = 16

    ws1.row_dimensions[1].height = 30
    ws1.merge_cells("A1:G1")
    t = ws1.cell(row=1, column=1, value="NailVesta 中台运营分析报告")
    t.font = Font(name="Arial", size=18, bold=True, color=WHITE)
    t.fill = fill(NAVY); t.alignment = center()

    ws1.merge_cells("A2:G2")
    s = ws1.cell(row=2, column=1, value=f"统计周期：{period_label}（本周，11天）｜ 对比：上周（7天）& 4月整月周均")
    s.font = Font(name="Arial", size=10, color="444444"); s.fill = fill(LGRAY); s.alignment = center()

    write_section_title(ws1, 4, 1, "一、核心经营指标三期对比", 7)
    hdrs = ["指标","4月周均（基准）","上周","vs4月","本周","WoW↑↓","本周vs4月"]
    write_header_row(ws1, 5, hdrs)

    kpi_rows = [
        ["Orders 订单量", f"{apr_orders:,.0f}（全渠道）", prev_orders, f"{pct(prev_orders,apr_orders):+.1f}%", cur_orders, f"{pct(cur_orders,prev_orders):+.1f}%", f"{pct(cur_orders,apr_orders):+.1f}%"],
        ["SKU sold 销量", f"{apr_sku:,.0f}（全渠道）", prev_sku, f"{pct(prev_sku,apr_sku):+.1f}%", cur_sku, f"{pct(cur_sku,prev_sku):+.1f}%", f"{pct(cur_sku,apr_sku):+.1f}%"],
        ["连带率", f"{apr_attach:.2f}x", f"{prev_attach:.2f}x", f"{pct(prev_attach,apr_attach):+.1f}%", f"{cur_attach:.2f}x", f"{pct(cur_attach,prev_attach):+.1f}%", f"{pct(cur_attach,apr_attach):+.1f}%"],
        ["AOV 订单均价($)", f"${apr_aov:.2f}", f"${prev_aov:.2f}", f"{pct(prev_aov,apr_aov):+.1f}%", f"${cur_aov:.2f}", f"{pct(cur_aov,prev_aov):+.1f}%", f"{pct(cur_aov,apr_aov):+.1f}%"],
        ["ASP 客单价($)", f"${apr_asp:.2f}", f"${prev_asp:.2f}", f"{pct(prev_asp,apr_asp):+.1f}%", f"${cur_asp:.2f}", f"{pct(cur_asp,prev_asp):+.1f}%", f"{pct(cur_asp,apr_asp):+.1f}%"],
        ["CTR 点击率", "—", f"{prev_ctr:.2f}%", "—", f"{cur_ctr:.2f}%", f"{pct(cur_ctr,prev_ctr):+.1f}%", "—"],
        ["CVR 转化率", "—", f"{prev_cvr:.2f}%", "—", f"{cur_cvr:.2f}%", f"{pct(cur_cvr,prev_cvr):+.1f}%", "—"],
        ["Add-to-Cart Rate", "—", f"{prev_atc:.2f}%", "—", f"{cur_atc:.2f}%", f"{pct(cur_atc,prev_atc):+.1f}%", "—"],
        ["总GMV($，直播)", f"${apr_gmv:,.0f}（全渠道）", f"${prev_gmv:,.0f}", f"{pct(prev_gmv,apr_gmv):+.1f}%", f"${cur_gmv:,.0f}", f"{pct(cur_gmv,prev_gmv):+.1f}%", f"{pct(cur_gmv,apr_gmv):+.1f}%"],
    ]
    for i, row_data in enumerate(kpi_rows):
        r = 6 + i
        alt = (i % 2 == 1)
        write_data_row(ws1, r, row_data, alt=alt)
        # Color WoW column
        wow_cell = ws1.cell(row=r, column=6)
        if wow_cell.value and "+" in str(wow_cell.value):
            wow_cell.font = Font(name="Arial", size=10, color="375623", bold=True)
        elif wow_cell.value and "-" in str(wow_cell.value):
            wow_cell.font = Font(name="Arial", size=10, color="7B0000", bold=True)

    # Analysis notes
    ws1.row_dimensions[16].height = 15
    write_section_title(ws1, 17, 1, "关键分析", 7, bg=BLUE)
    notes = [
        ["量级指标", "订单量 ↓17.1%（1,580→1,310），GMV ↓13.2%（$65K→$56K）两连跌。含劳动节假期影响，连带率 1.47x 从低点修复（↑2.8%），量级下滑主要来自流量端。"],
        ["质量指标", f"AOV ↑4.6%（${prev_aov}→${cur_aov}）、ASP ↑2.1%（${prev_asp}→${cur_asp}）同步回升，ASP 基本追平 4 月月均（${apr_asp}）。CVR 突破 1%（↑7.4%）。"],
        ["最大预警", "Add-to-Cart Rate ↓39.6%（7.38%→4.46%），ATC 低位若持续将在下周直接压制 GMV 天花板，需立即排查直播素材与商品卡。"],
    ]
    for i, (cat, note) in enumerate(notes):
        r = 18 + i
        ws1.row_dimensions[r].height = 35
        c1 = ws1.cell(row=r, column=1, value=cat)
        c1.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c1.fill = fill(BLUE); c1.alignment = center(); c1.border = border()
        c2 = ws1.cell(row=r, column=2, value=note)
        c2.font = cell_font(size=10); c2.fill = fill(LGRAY if i%2==0 else WHITE)
        c2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); c2.border = border()
        ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)

    # ── SHEET 2: SKU 结构 ──────────────────────────────
    ws2 = wb.create_sheet("SKU结构与AOV")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 10
    for col in ["B","C","D","E","F","G","H"]:
        ws2.column_dimensions[col].width = 15

    ws2.merge_cells("A1:H1")
    t2 = ws2.cell(row=1, column=1, value="二、订单 SKU 结构 & AOV 分层")
    t2.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t2.fill = fill(NAVY); t2.alignment = center()

    write_header_row(ws2, 2, ["SKU/单","本周订单数","本周占比","本周AOV","上周占比","上周AOV","4月占比","4月AOV"])
    apr_pcts = [73.8,15.7,2.5,None,8.1]
    apr_aovs = [34.16,59.93,94.40,None,130.11]
    for i, r in enumerate(sku_rows):
        row_n = 3 + i
        alt = (i % 2 == 1)
        data = [
            r["tier"], r["count"], f'{r["pct"]:.1f}%', f'${r["aov"]:.2f}',
            f'{r["p_pct"]:.1f}%', f'${r["p_aov"]:.2f}',
            f'{apr_pcts[i]:.1f}%' if apr_pcts[i] else "—",
            f'${apr_aovs[i]:.2f}' if apr_aovs[i] else "—",
        ]
        write_data_row(ws2, row_n, data, alt=alt)

    ws2.row_dimensions[9].height = 15
    write_section_title(ws2, 10, 1, "结构分析", 8, bg=BLUE)
    struct_notes = [
        "✅ 1件单占比从上周高点回落（76.1%→73.6%），是 AOV/ASP 同步回升的核心驱动。",
        "⚠️ 2件单 AOV $55.52 仍低于满减门槛 $60，活动2对该档客群激励有限（4月月报门槛错配结论延续）。",
        "🌟 4+件 AOV $197.55 创近期新高（4月$130→上周$183→本周$197），高价值客群消费意愿极强，建议设计专属激励路径。",
    ]
    for i, note in enumerate(struct_notes):
        r = 11 + i
        ws2.row_dimensions[r].height = 30
        c = ws2.cell(row=r, column=1, value=note)
        c.font = cell_font(size=10); c.fill = fill(LGRAY if i%2==0 else WHITE)
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True); c.border = border()
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

    # ── SHEET 3: 退换货 ──────────────────────────────
    ws3 = wb.create_sheet("退换货分析")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["A"].width = 28
    for col in ["B","C","D","E"]:
        ws3.column_dimensions[col].width = 16

    ws3.merge_cells("A1:E1")
    t3 = ws3.cell(row=1, column=1, value="三、退换货健康度分析")
    t3.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t3.fill = fill(NAVY); t3.alignment = center()

    # Summary metrics
    ret_summary = [
        ("总退货率", f"{ret_rate:.2f}%", f"4月月均 {apr_ret_rate:.2f}%，↓{abs(pct(ret_rate,apr_ret_rate)):.1f}%", "375623"),
        ("退货包裹（全渠道）", f"{returned_pkgs}", f"11天 · 周均≈{int(returned_pkgs/11*7)}", "1F3864"),
        ("Per Order Return", f"${per_order_ret:.2f}", "4月 $46.33 · ↓16.1%", "375623"),
        ("Seller Fault", f"{seller_fault}单", "占退货 48.5% · 仓库预警", "7B0000"),
        ("取消订单（全渠道）", f"{cancel_cnt}", f"取消率 {cancel_cnt/total_orders_full*100:.1f}%", "7F6000"),
    ]
    for i, (lbl, val, sub, color) in enumerate(ret_summary):
        write_metric_block(ws3, 3, 1+i, lbl, val, sub, color)
        ws3.column_dimensions[get_column_letter(1+i)].width = 18

    write_section_title(ws3, 7, 1, "退货原因构成", 5, bg=RED)
    write_header_row(ws3, 8, ["退货原因","数量","占比","性质","趋势"], bg="C00000")
    ret_data = [
        ["No longer needed",100,"51.5%","非卖家","—"],
        ["Missing package",20,"10.3%","物流","—"],
        ["Item doesn't match desc.",14,"7.2%","卖家描述","持续"],
        ["Wrong item was sent",14,"7.2%","仓库","↑160% 🔴"],
        ["Damaged item/packaging",13,"6.7%","仓库","↑400% 🔴"],
        ["Missing items",13,"6.7%","仓库","↑200% 🔴"],
        ["Item arrived too late",6,"3.1%","物流","—"],
        ["Defective item",5,"2.6%","产品质量","—"],
        ["Missing parts",4,"2.1%","仓库","—"],
    ]
    for i, row_d in enumerate(ret_data):
        write_data_row(ws3, 9+i, row_d, alt=(i%2==1))
        # Highlight warehouse faults
        if "↑" in str(row_d[4]):
            for col in range(1, 6):
                ws3.cell(row=9+i, column=col).fill = fill(LRED)
                ws3.cell(row=9+i, column=col).font = Font(name="Arial", size=10, color=RED)

    write_section_title(ws3, 19, 1, "高退货产品 & 款式 Top7", 5, bg=RED)
    write_header_row(ws3, 20, ["产品/款式","退货包裹","占比","风险级别","备注"], bg="C00000")
    top_ret_data = [
        ["DreamWear Collection",38,"18.4%","🔴 最高危","优先排查库存映射与拣货标签"],
        ["Next Gen Collection",15,"7.8%","🟡 中","—"],
        ["NEW DROP Collection",15,"6.9%","🟡 中","—"],
        ["Organizer Binder（非甲片）",10,"5.5%","🔴 描述偏差","检查详情页与实物一致性"],
        ["Acai Bloom",9,"4.1%","🟡 延续","—"],
        ["Prism Aura / Ruby Bloom",9,"4.1%","🟡 延续","—"],
        ["Rosé Petal",8,"3.7%","🔴 跨月持续","4月月报最高危款，立即供应商复盘"],
    ]
    for i, row_d in enumerate(top_ret_data):
        write_data_row(ws3, 21+i, row_d, alt=(i%2==1))

    write_section_title(ws3, 29, 1, "取消原因分布（全渠道 103单）", 5, bg=AMBER)
    write_header_row(ws3, 30, ["取消原因","订单数","占比","性质","备注"], bg="BF8F00")
    cancel_data = [
        ["Bought by mistake",45,"43.7%","冲动下单反悔","直播引导强但确认度低"],
        ["No longer needed",28,"27.2%","主动取消","—"],
        ["Incorrect shipping address",9,"8.7%","操作失误","—"],
        ["Need to change payment",6,"5.8%","支付问题","—"],
        ["Customer overdue to pay",5,"4.9%","超时系统取消","—"],
        ["Discount not as expected",4,"3.9%","优惠预期差","活动2传达问题"],
        ["Forgot to apply coupons",3,"2.9%","操作失误","结账页提示未落地"],
    ]
    for i, row_d in enumerate(cancel_data):
        write_data_row(ws3, 31+i, row_d, alt=(i%2==1))

    # ── SHEET 4: Auction ──────────────────────────────
    ws4 = wb.create_sheet("Auction专项")
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions["A"].width = 22
    for col in ["B","C","D","E"]:
        ws4.column_dimensions[col].width = 18

    ws4.merge_cells("A1:E1")
    t4 = ws4.cell(row=1, column=1, value="四、Auction 渠道专项分析")
    t4.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t4.fill = fill(NAVY); t4.alignment = center()

    auc_sum = [
        ("总Auction订单", f"{auc_total}", f"全渠道{total_orders_full}单的{auc_total/total_orders_full*100:.1f}%","1F3864"),
        ("Auction退货率", f"{auc_ret_rate}%", f"全渠道{ret_rate:.2f}%，低{abs(pct(auc_ret_rate,ret_rate)):.0f}%","375623"),
        ("Auction取消率", f"{auc_cancel_rate}%", f"{auc_cancel}单·100%超时未付","7F6000"),
        ("有效Auction AOV", f"${auc_aov:.2f}", f"排除{auc_cancel}单取消","1F3864"),
        ("退货Auction AOV", f"${auc_ret_aov:.2f}", f"高于均值+${auc_ret_aov-auc_aov:.2f}","7F6000"),
    ]
    for i, (lbl,val,sub,color) in enumerate(auc_sum):
        write_metric_block(ws4, 3, 1+i, lbl, val, sub, color)

    write_section_title(ws4, 7, 1, "Auction vs 全渠道健康度对比", 5, bg=GREEN)
    write_header_row(ws4, 8, ["指标","Auction渠道","全渠道","差异","评价"], bg="375623")
    cmp_data = [
        ["退货率", f"{auc_ret_rate}%", f"{ret_rate:.2f}%", f"低{abs(pct(auc_ret_rate,ret_rate)):.0f}%","✅ 显著优于全渠道"],
        ["取消率", f"{auc_cancel_rate}%（全为超时）", "4.0%（多种原因）","—","⚠️ 催付机制可解决"],
        ["Seller Fault占退货","25.0%","48.5%","↓48.5%","✅ 仓库错包未扩散"],
        ["退货主因","No longer needed 75%","No longer needed 51.5%","—","✅ 健康"],
        ["仓库错包渗透","极低","Wrong/Missing激增","—","✅ Auction未受影响"],
        ["AOV",f"${auc_aov:.2f}",f"${cur_aov:.2f}",f"${cur_aov-auc_aov:.2f}","参考"],
    ]
    for i, row_d in enumerate(cmp_data):
        write_data_row(ws4, 9+i, row_d, alt=(i%2==1))
        if "✅" in str(row_d[4]):
            ws4.cell(row=9+i, column=5).font = Font(name="Arial", size=10, color=GREEN)
        elif "⚠️" in str(row_d[4]):
            ws4.cell(row=9+i, column=5).font = Font(name="Arial", size=10, color=AMBER)

    write_section_title(ws4, 16, 1, "Auction SKU 销售 vs 取消", 5, bg=BLUE)
    write_header_row(ws4, 17, ["SKU","销售占比","取消占比","隐含取消率","评价"])
    auc_sku_data = [
        ["SKU 1","51.8%","31.2%","3.4%","✅ 正常"],
        ["SKU 2","32.7%","50.0%","8.6%","🔴 超比例取消"],
        ["SKU 3","13.7%","12.5%","6.4%","正常"],
        ["SKU 4","1.8%","6.2%","—","占比极小"],
    ]
    for i, row_d in enumerate(auc_sku_data):
        write_data_row(ws4, 18+i, row_d, alt=(i%2==1))
        if "🔴" in str(row_d[4]):
            for col in range(1,6):
                ws4.cell(row=18+i, column=col).fill = fill(LRED)

    write_section_title(ws4, 23, 1, "Auction 退货原因 & 取消产品", 5, bg=BLUE)
    write_header_row(ws4, 24, ["退货原因","数量","占比","取消产品","取消数"])
    auc_det = [
        ["No longer needed",9,"75.0%","Choose 2 Sets",12],
        ["Missing package",3,"25.0%","Pick your two sets and size",2],
        ["—","—","—","其他",2],
    ]
    for i, row_d in enumerate(auc_det):
        write_data_row(ws4, 25+i, row_d, alt=(i%2==1))

    ws4.cell(row=29, column=1, value="核心结论：Auction 是本期最健康的子渠道。退货率 2.5% 仅为全渠道 44%，Seller Fault 25% vs 全渠道 48.5%。唯一问题：16 单超时取消——建立直播后 30 分钟催付提醒可零成本挽回，并针对 SKU 2（取消率 8.6%）优化直播话术。")
    ws4.cell(row=29, column=1).font = Font(name="Arial", size=10, italic=True, color="444444")
    ws4.cell(row=29, column=1).fill = fill(LGREEN)
    ws4.cell(row=29, column=1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws4.row_dimensions[29].height = 45
    ws4.merge_cells("A29:E29")

    # ── SHEET 5: 行动建议 ──────────────────────────────
    ws5 = wb.create_sheet("行动建议")
    ws5.sheet_view.showGridLines = False
    ws5.column_dimensions["A"].width = 8
    ws5.column_dimensions["B"].width = 25
    ws5.column_dimensions["C"].width = 55
    ws5.column_dimensions["D"].width = 18

    ws5.merge_cells("A1:D1")
    t5 = ws5.cell(row=1, column=1, value="五、行动建议（衔接 4 月月报）")
    t5.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    t5.fill = fill(NAVY); t5.alignment = center()

    write_header_row(ws5, 2, ["#","行动项","具体内容","优先级"], bg=BLUE)
    actions = [
        (1,"仓库批次核查","Wrong item ↑160% / Missing ↑200% / Damaged ↑400%，回溯 5/1–5/11 出库批次，锁定拣货问题节点。4 月月报已预警 W4 后错包率升至 5%，本期是直接延续，不立即切断将影响退货率改善成果。","🔴 立即"),
        (2,"DreamWear Listing 专查","退货 38 包裹占全渠道 18.4%，核查库存映射与主图实物一致性，结合 Wrong item 比例判断是否为拣货标签错误。","🔴 立即"),
        (3,"Rosé Petal 供应商复盘","4 月月报已标注为持续最高危款，本期再入 Top。提取退货评论，决定是否暂停销售或全线整改，不得再拖。","🔴 立即"),
        (4,"Auction 催付机制","16 单 100% 超时取消，直播结束后 30 分钟内建立催付提醒，零成本挽回，重点针对 SKU 2（取消率 8.6%）。","🔴 立即"),
        (5,"Add-to-Cart 归因分析","↓39.6% 是最大流量预警。ATC 低位若延续将在下周直接压制 GMV，复盘直播素材、封面与商品卡点击路径。","🟡 本月"),
        (6,"结账页 1→2件转化提示","对 AOV $30-$55 用户推再加一件立减$8，4 月月报已建议，尚未落地，是 AOV / ASP 持续提升最直接杠杆。","🟡 本月"),
        (7,"Organizer Binder 描述核查","非甲片类产品退货 10 包裹，检查详情页内容物描述与实物一致性，Missing items 是主要退货原因。","🟡 本月"),
        (8,"Auction SKU 2 复查","取消率 8.6% vs SKU 1 的 3.4%，检查直播中 SKU 2 定价展示与买家预期一致性，优化确认话术。","🟡 本月"),
        (9,"活动 2 发券精准化","活动 2 延续至 5/4，4 月月报指出约 80% 领券为无效领取，将领券资格限定为历史多件购买用户，目标使用率从 9% 提升至 30%+，ROI 从 10.97x 恢复至 15x+。","🟢 策略层"),
        (10,"非直播 Listing 优化","非直播取消率 5.7% 持续高于直播 3.3%，完善主图尺码信息与详情页描述，减少搜索流量预期差。","🟢 策略层"),
        (11,"统一 GMV 统计口径","直播口径与全渠道口径混用影响月度趋势判断，本月内确定统一标准并追溯历史数据。","🟢 策略层"),
        (12,"连带率 & 错包率预警线监控","连带率预警线 1.5x（本期 1.47x 已低于此），退换货错包率预警线 2%（本期已突破）。下周需重点监控两项指标。","🟢 策略层"),
    ]
    priority_fill = {"🔴 立即": LRED, "🟡 本月": LAMBER, "🟢 策略层": LGREEN}
    priority_color = {"🔴 立即": RED, "🟡 本月": AMBER, "🟢 策略层": GREEN}
    for i, (num, title, detail, pri) in enumerate(actions):
        r = 3 + i
        ws5.row_dimensions[r].height = 45
        bg = priority_fill.get(pri, WHITE)
        ws5.cell(row=r, column=1, value=num).font = Font(name="Arial", size=11, bold=True, color=WHITE)
        ws5.cell(row=r, column=1).fill = fill(priority_color.get(pri,"1F3864").replace("🔴","").replace("🟡","").replace("🟢",""))
        ws5.cell(row=r, column=1).alignment = center(); ws5.cell(row=r, column=1).border = border()
        for col, val in [(2,title),(3,detail),(4,pri)]:
            c = ws5.cell(row=r, column=col, value=val)
            c.font = Font(name="Arial", size=10, bold=(col==2))
            c.fill = fill(bg); c.alignment = Alignment(horizontal="left",vertical="center",wrap_text=True)
            c.border = border()
        ws5.cell(row=r, column=1).fill = fill(priority_color.get(pri,"1F3864"))

    # Fix action number cell colors
    for i, (num, title, detail, pri) in enumerate(actions):
        r = 3 + i
        c = ws5.cell(row=r, column=1)
        c.fill = fill(priority_color.get(pri,"1F3864"))
        c.font = Font(name="Arial", size=11, bold=True, color=WHITE)

    # Final formatting: freeze panes on all sheets
    for ws in [ws1,ws2,ws3,ws4,ws5]:
        ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ══════════════════════════════════════════════
# DOWNLOAD BUTTON
# ══════════════════════════════════════════════
st.markdown("---")
st.markdown("### 📥 下载 Excel 报告")

if st.button("生成并下载 Excel 报告 ⬇️", type="primary"):
    with st.spinner("正在生成 Excel 报告..."):
        excel_buf = build_excel()
        st.download_button(
            label="⬇️ 下载 NailVesta_周报.xlsx",
            data=excel_buf,
            file_name=f"NailVesta_周报_{period_label.replace('/','').replace('–','_')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    st.success("✅ Excel 报告已生成！点击上方按钮下载。")

st.markdown("---")
st.caption("NailVesta 中台运营分析系统 · 数据源：Auction Report / Cancelled Report / Returned Report / 运营截图")
